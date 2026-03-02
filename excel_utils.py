from openpyxl import Workbook, load_workbook
from io import BytesIO
import re

EXCEL_HEADERS = [
    'tipo_documento', 'numero_documento', 'primer_nombre', 'segundo_nombre',
    'primer_apellido', 'segundo_apellido', 'correo_institucional', 'correo_personal',
    'celular', 'direccion', 'programa'
]


def export_students_to_excel(students):
    wb = Workbook()
    ws = wb.active
    ws.append(EXCEL_HEADERS)
    for s in students:
        row = [
            s.tipo_documento or '',
            s.numero_documento or '',
            s.primer_nombre or '',
            s.segundo_nombre or '',
            s.primer_apellido or '',
            s.segundo_apellido or '',
            s.correo_institucional or '',
            s.correo_personal or '',
            s.celular or '',
            s.direccion or '',
            s.programa or ''
        ]
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_students_from_excel(file_stream):
    # file_stream: file-like object
    wb = load_workbook(filename=file_stream, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ['El archivo está vacío']
    headers = [h for h in rows[0]]
    headers = [str(h).strip() if h is not None else '' for h in headers]
    header_errors = []
    # comprobar cabeceras requeridas
    required = ['numero_documento', 'primer_nombre', 'primer_apellido']
    for req in required:
        if req not in headers:
            header_errors.append(f'Falta columna requerida: {req}')
    if header_errors:
        return [], header_errors, []

    results = []
    row_errors = []
    email_re = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    phone_re = re.compile(r"^[0-9+\-()\s]+$")
    for idx, r in enumerate(rows[1:], start=2):
        if not any(r):
            continue
        data = {}
        msgs = []
        for i, h in enumerate(headers):
            if h:
                data[h] = r[i] if i < len(r) else None

        numero = str(data.get('numero_documento') or '').strip()
        if not numero:
            msgs.append('Falta numero_documento')
        if not (data.get('primer_nombre')):
            msgs.append('Falta primer_nombre')
        if not (data.get('primer_apellido')):
            msgs.append('Falta primer_apellido')

        # validar emails
        for email_field in ('correo_institucional', 'correo_personal'):
            val = data.get(email_field)
            if val:
                if not email_re.match(str(val).strip()):
                    msgs.append(f'Formato inválido en {email_field}: {val}')

        # validar celular (solo caracteres válidos y longitud razonable)
        celular = data.get('celular')
        if celular:
            if not phone_re.match(str(celular).strip()):
                msgs.append(f'Formato inválido en celular: {celular}')
            else:
                digits = re.sub(r'\D', '', str(celular))
                if len(digits) < 6 or len(digits) > 15:
                    msgs.append(f'Celular con longitud inválida: {celular}')

        if msgs:
            row_errors.append({'row': idx, 'messages': msgs, 'data': data})
        results.append(data)

    return results, [], row_errors


def generate_template():
    wb = Workbook()
    ws = wb.active
    ws.append(EXCEL_HEADERS)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
